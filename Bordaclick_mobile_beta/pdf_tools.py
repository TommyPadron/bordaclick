import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

DATABASE = "bordaclick_dev.db"

def _safe_float(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def generar_pdf_orden(orden, detalle_orden, tasa_cambio=0.0):
    pedido_id = int(_safe_float(orden.get("id")))
    nombre_pdf = f"Pedido_{pedido_id:04d}.pdf"

    doc = SimpleDocTemplate(
        nombre_pdf,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    estilos = getSampleStyleSheet()
    
    estilo_titulo = ParagraphStyle(
        'TituloComprobante',
        parent=estilos['Normal'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#1D3557")
    )
    
    estilo_seccion = ParagraphStyle(
        'SeccionHeader',
        parent=estilos['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=colors.black
    )

    estilo_celda = ParagraphStyle('CeldaTabla', parent=estilos['Normal'], fontSize=9, leading=11)
    estilo_celda_centro = ParagraphStyle('CeldaTablaCentro', parent=estilos['Normal'], fontSize=9, leading=11, alignment=1)
    estilo_celda_header = ParagraphStyle('CeldaHeader', parent=estilos['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.white, alignment=1)

    elementos = []

    # 1. Encabezado Título##
    elementos.append(Paragraph(f"BORDACLICK - COMPROBANTE DE ORDEN #{pedido_id:04d}", estilo_titulo))
    elementos.append(Spacer(1, 15))

    # 2. Cuadro Datos Cliente / Entrega
    delivery_str = str(orden.get("delivery", "No"))
    if delivery_str == "Sí" and orden.get("zona_delivery"):
        delivery_str += f" ({orden.get('zona_delivery')})"

    datos_cuadro = [
        [
            Paragraph(f"<b>Cliente:</b> {orden.get('nombre', '')}", estilo_celda),
            Paragraph(f"<b>Fecha Entrega:</b> {orden.get('fecha_entrega', '')}", estilo_celda)
        ],
        [
            Paragraph(f"<b>Teléfono:</b> {orden.get('telefono', '')}", estilo_celda),
            Paragraph(f"<b>Estado:</b> {orden.get('status', '')}", estilo_celda)
        ],
        [
            Paragraph(f"<b>Correo:</b> {orden.get('correo', '')}", estilo_celda),
            Paragraph(f"<b>Delivery:</b> {delivery_str}", estilo_celda)
        ]
    ]

    tabla_info = Table(datos_cuadro, colWidths=[260, 260])
    tabla_info.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8F9FA")),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#CED4DA")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E9ECEF")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
    ]))
    elementos.append(tabla_info)
    elementos.append(Spacer(1, 20))

    # 3. Detalle de Prendas
    elementos.append(Paragraph("DETALLE DE PRENDAS", estilo_seccion))
    elementos.append(Spacer(1, 8))

    tabla_prendas_datos = [[
        Paragraph("Colegio", estilo_celda_header),
        Paragraph("Prenda", estilo_celda_header),
        Paragraph("Talla", estilo_celda_header),
        Paragraph("Marca", estilo_celda_header),
        Paragraph("Color", estilo_celda_header),
        Paragraph("Cant.", estilo_celda_header)
    ]]

    if not detalle_orden.empty:
        for _, fila in detalle_orden.iterrows():
            tabla_prendas_datos.append([
                Paragraph(str(fila.get("Colegio", "")), estilo_celda),
                Paragraph(str(fila.get("Tipo Prenda", "")), estilo_celda_centro),
                Paragraph(str(fila.get("Talla", "")), estilo_celda_centro),
                Paragraph(str(fila.get("Marca", "")), estilo_celda_centro),
                Paragraph(str(fila.get("Color", "")), estilo_celda_centro),
                Paragraph(str(fila.get("Cantidad", "")), estilo_celda_centro)
            ])

    tabla_prendas = Table(tabla_prendas_datos, colWidths=[140, 100, 60, 90, 80, 50])
    tabla_prendas.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1D3557")),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#DEE2E6")),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elementos.append(tabla_prendas)
    elementos.append(Spacer(1, 15))

    # 4. Resumen Financiero
    sub_bordado = _safe_float(orden.get("subtotal_bordado"))
    sub_nombres = _safe_float(orden.get("subtotal_nombres"))
    del_costo = _safe_float(orden.get("delivery_costo"))
    abono = _safe_float(orden.get("abono"))
    saldo = _safe_float(orden.get("saldo_pendiente"))
    total_general = sub_bordado + sub_nombres + del_costo

    estilo_fin_lbl = ParagraphStyle('FinLbl', parent=estilos['Normal'], fontName='Helvetica', fontSize=10, alignment=2)
    estilo_fin_val = ParagraphStyle('FinVal', parent=estilos['Normal'], fontName='Helvetica', fontSize=10, alignment=2)
    estilo_fin_bold_lbl = ParagraphStyle('FinBoldLbl', parent=estilos['Normal'], fontName='Helvetica-Bold', fontSize=10, alignment=2)
    estilo_fin_bold_val = ParagraphStyle('FinBoldVal', parent=estilos['Normal'], fontName='Helvetica-Bold', fontSize=10, alignment=2)

    tabla_fin_datos = [
        [Paragraph("Subtotal Bordados:", estilo_fin_lbl), Paragraph(f"${sub_bordado:.2f}", estilo_fin_val)],
        [Paragraph("Subtotal Nombres:", estilo_fin_lbl), Paragraph(f"${sub_nombres:.2f}", estilo_fin_val)],
        [Paragraph("Costo Delivery:", estilo_fin_lbl), Paragraph(f"${del_costo:.2f}", estilo_fin_val)],
        [Paragraph("TOTAL GENERAL:", estilo_fin_bold_lbl), Paragraph(f"${total_general:.2f}", estilo_fin_bold_val)],
        [Paragraph("Abonado:", estilo_fin_lbl), Paragraph(f"${abono:.2f}", estilo_fin_val)],
        [Paragraph("Saldo Pendiente:", estilo_fin_bold_lbl), Paragraph(f"${saldo:.2f}", estilo_fin_bold_val)]
    ]

    if tasa_cambio > 0:
        saldo_bs = saldo * tasa_cambio
        total_bs = total_general * tasa_cambio
        tabla_fin_datos.append([Paragraph("Tasa BCV/Cambio:", estilo_fin_lbl), Paragraph(f"{tasa_cambio:.2f} Bs.", estilo_fin_val)])
        tabla_fin_datos.append([Paragraph("Total en Bs.:", estilo_fin_bold_lbl), Paragraph(f"{total_bs:,.2f} Bs.", estilo_fin_bold_val)])
        tabla_fin_datos.append([Paragraph("Saldo Pendiente en Bs.:", estilo_fin_bold_lbl), Paragraph(f"{saldo_bs:,.2f} Bs.", estilo_fin_bold_val)])

    tabla_fin = Table(tabla_fin_datos, colWidths=[150, 100])
    tabla_fin.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('LINEABOVE', (0,3), (1,3), 1, colors.black),
    ]))

    tabla_contenedor = Table([[ "", tabla_fin ]], colWidths=[270, 250])
    tabla_contenedor.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))

    elementos.append(tabla_contenedor)

    doc.build(elementos)
    return nombre_pdf


def generar_excel_orden(orden, detalle_orden):
    pedido_id = int(_safe_float(orden.get("id")))
    nombre_excel = f"Pedido_{pedido_id:04d}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Orden_Servicio"

    titulo_blanco = Font(bold=True, color="FFFFFF", size=12)
    subtitulo_bold = Font(bold=True, size=11)
    encabezado_tabla = Font(bold=True, color="FFFFFF")
    
    fondo_azul_oscuro = PatternFill(fill_type="solid", start_color="1A365D")
    fondo_azul_medio = PatternFill(fill_type="solid", start_color="2B6CB0")

    borde = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")

    ws["A1"] = "BORDACLICK"
    ws["A2"] = "ORDEN DE SERVICIO"

    for cell in [ws["A1"], ws["A2"]]:
        cell.font = titulo_blanco
        cell.fill = fondo_azul_oscuro
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fecha_pago_val = str(orden.get("fecha_pago", "Sin registro")) if orden.get("fecha_pago") else "Sin registro"

    datos = [
        ("A4", "Pedido", f"#{pedido_id:04d}"),
        ("A5", "Estado", orden.get("status", "")),
        ("A6", "Fecha Entrega", str(orden.get("fecha_entrega", ""))),
        ("A7", "Fecha Último Pago", fecha_pago_val),
        ("A9", "DATOS DEL CLIENTE", ""),
        ("A10", "Nombre", orden.get("nombre", "")),
        ("A11", "Teléfono", orden.get("telefono", "")),
        ("A12", "Correo", orden.get("correo", "")),
        ("A13", "Colegio", orden.get("colegio", "")),
        ("A15", "DATOS DE PRODUCCIÓN", ""),
        ("A16", "Tipo Logo", orden.get("tipo_logo", "")),
        ("A17", "Cantidad Total", orden.get("cantidad_total", 0)),
        ("A18", "Nombre Bordado", orden.get("nombre_bordado", "N/A")),
        ("A19", "Cantidad con Nombre", orden.get("cantidad_nombre", 0)),
        ("A21", "RESUMEN FINANCIERO", ""),
        ("A22", "Subtotal Bordado", _safe_float(orden.get("subtotal_bordado"))),
        ("A23", "Subtotal Nombres", _safe_float(orden.get("subtotal_nombres"))),
        ("A24", "Delivery", _safe_float(orden.get("delivery_costo"))),
        ("A25", "Abono USD", _safe_float(orden.get("abono"))),
        ("A26", "Saldo Pendiente USD", _safe_float(orden.get("saldo_pendiente")))
    ]

    for c_pos, label, val in datos:
        ws[c_pos] = label
        row_num = int(c_pos[1:])
        
        if label in ["DATOS DEL CLIENTE", "DATOS DE PRODUCCIÓN", "RESUMEN FINANCIERO"]:
            ws[c_pos].font = subtitulo_bold
        else:
            ws[f"B{row_num}"] = val
            ws[c_pos].font = Font(bold=True)
            if isinstance(val, float):
                ws[f"B{row_num}"].number_format = "$#,##0.00"

    fila_excel = 28
    ws[f"A{fila_excel}"] = "DESGLOSE DE PRENDAS"
    ws[f"A{fila_excel}"].font = subtitulo_bold
    fila_excel += 1

    inicio_desglose = fila_excel

    if not detalle_orden.empty and "Colegio" in detalle_orden.columns:
        for colegio in detalle_orden["Colegio"].unique():
            ws[f"A{fila_excel}"] = f"🏫 {colegio}"
            ws[f"A{fila_excel}"].font = subtitulo_bold
            fila_excel += 1

            headers = ["Tipo Prenda", "Talla", "Marca", "Color", "Cantidad"]
            cols = ["A", "B", "C", "D", "E"]

            for col_letter, h_text in zip(cols, headers):
                cell = ws[f"{col_letter}{fila_excel}"]
                cell.value = h_text
                cell.font = encabezado_tabla
                cell.fill = fondo_azul_medio
                cell.alignment = Alignment(horizontal="center")

            fila_excel += 1

            df_colegio = detalle_orden[detalle_orden["Colegio"] == colegio]
            for _, fila in df_colegio.iterrows():
                ws[f"A{fila_excel}"] = fila.get("Tipo Prenda", "")
                ws[f"B{fila_excel}"] = fila.get("Talla", "")
                ws[f"C{fila_excel}"] = fila.get("Marca", "")
                ws[f"D{fila_excel}"] = fila.get("Color", "")
                ws[f"E{fila_excel}"] = fila.get("Cantidad", 0)
                fila_excel += 1

            fila_excel += 1

    for fila in ws.iter_rows(min_row=inicio_desglose, max_row=fila_excel - 1, min_col=1, max_col=5):
        for celda in fila:
            if celda.value:
                celda.border = borde

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15

    wb.save(nombre_excel)
    return nombre_excel


def generar_excel_historico(df_ordenes):
    nombre_excel = "Historico_Ordenes_Bordaclick.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Historico_Ordenes"

    titulo_blanco = Font(bold=True, color="FFFFFF", size=12)
    encabezado_tabla = Font(bold=True, color="FFFFFF")
    fondo_azul_oscuro = PatternFill(fill_type="solid", start_color="1A365D")
    fondo_azul_medio = PatternFill(fill_type="solid", start_color="2B6CB0")

    ws.merge_cells("A1:N1")
    ws["A1"] = "HISTÓRICO GENERAL DE ÓRDENES Y PAGOS EN BOLÍVARES Y DÓLARES - BORDACLICK"
    ws["A1"].font = titulo_blanco
    ws["A1"].fill = fondo_azul_oscuro
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "ID Pedido", "Cliente", "Teléfono", "Correo", "Colegio",
        "Cant. Prendas", "Delivery", "Estado", "Fecha Entrega",
        "Abonado ($)", "Saldo Pendiente ($)", "Monto Pago (Bs.)", "Tasa de Cambio", "Fecha Último Pago"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = encabezado_tabla
        cell.fill = fondo_azul_medio
        cell.alignment = Alignment(horizontal="center")

    conn = sqlite3.connect(DATABASE)

    for row_idx, fila in enumerate(df_ordenes.to_dict(orient="records"), start=4):
        orden_id = int(fila.get('id', 0))
        ws.cell(row=row_idx, column=1, value=f"#{orden_id:04d}")
        ws.cell(row=row_idx, column=2, value=str(fila.get("nombre", "")))
        ws.cell(row=row_idx, column=3, value=str(fila.get("telefono", "")))
        ws.cell(row=row_idx, column=4, value=str(fila.get("correo", "")))
        ws.cell(row=row_idx, column=5, value=str(fila.get("colegio", "")))
        ws.cell(row=row_idx, column=6, value=int(fila.get("cantidad_total", 0)))
        ws.cell(row=row_idx, column=7, value=str(fila.get("delivery", "")))
        ws.cell(row=row_idx, column=8, value=str(fila.get("status", "")))
        ws.cell(row=row_idx, column=9, value=str(fila.get("fecha_entrega", "")))

        celda_abono = ws.cell(row=row_idx, column=10, value=_safe_float(fila.get("abono")))
        celda_abono.number_format = "$#,##0.00"

        celda_saldo = ws.cell(row=row_idx, column=11, value=_safe_float(fila.get("saldo_pendiente")))
        celda_saldo.number_format = "$#,##0.00"

        # Cargar los datos específicos del pago registrado en Bolívares
        df_pago = pd.read_sql_query(f"SELECT monto_bs, tasa_cambio FROM historico_pagos WHERE orden_id = {orden_id} ORDER BY id DESC LIMIT 1", conn)
        
        monto_bs_val = df_pago["monto_bs"].iloc[0] if not df_pago.empty else 0.0
        tasa_val = df_pago["tasa_cambio"].iloc[0] if not df_pago.empty else 0.0

        celda_bs = ws.cell(row=row_idx, column=12, value=_safe_float(monto_bs_val))
        celda_bs.number_format = "Bs.#,##0.00"

        celda_tasa = ws.cell(row=row_idx, column=13, value=_safe_float(tasa_val))
        celda_tasa.number_format = "#,##0.00"

        fecha_pago_val = str(fila.get("fecha_pago", "Sin pagos")) if fila.get("fecha_pago") else "Sin pagos"
        ws.cell(row=row_idx, column=14, value=fecha_pago_val)

    conn.close()

    columnas = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
    anchos = [12, 25, 15, 25, 25, 14, 12, 18, 15, 15, 15, 18, 15, 18]
    for col, ancho in zip(columnas, anchos):
        ws.column_dimensions[col].width = ancho

    wb.save(nombre_excel)
    return nombre_excel