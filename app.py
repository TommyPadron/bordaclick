import streamlit as st
import re
import pandas as pd
import smtplib

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side


from openpyxl import Workbook
from datetime import date,timedelta
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    PageBreak,
    Table,
    TableStyle,
    Image
)
from reportlab.lib.styles import getSampleStyleSheet
import os

from database import (
    crear_bd,
    guardar_orden,
    guardar_detalle,
    guardar_precio_bordado,
    obtener_precio_bordado,
    obtener_catalogo_bordados,
    guardar_catalogo_bordados,
    guardar_parametro,
    obtener_parametro,
    obtener_ordenes,
    obtener_orden_por_id,
    obtener_detalle_orden,
    actualizar_status_orden,
    guardar_colegio,
    obtener_colegios,
    obtener_precio_colegio,
    guardar_tipo_prenda,
    obtener_tipos_prenda,
    guardar_marca,
    obtener_marcas,
    guardar_color,
    obtener_colores,
    guardar_talla,
    obtener_tallas,
    guardar_color,
    obtener_colores,
    registrar_pago,
    enviar_pdf_por_correo,
    guardar_zona_delivery,
    obtener_zonas_delivery,
    obtener_costo_delivery,
    enviar_notificacion_estado,
    contar_pedidos_pendientes,
)          

st.set_page_config(
    page_title="Bordaclick",
    page_icon="🧵",
    layout="wide"
)

crear_bd()

st.image(

    "Logo Bordaclick.JPG",
    width=250
)


colegios = [
    "San Ignacio",
    "Colegio Don Bosco - Altamira",
    "Colegio Simón Bolívar",
    "IEA",
    "Colegio Valle Abierto",
    "Colegio Santiago de León"
]

clave_admin = st.sidebar.text_input(
    "Clave Administrador",
    type="password"
)

#st.write(repr(clave_admin))

opciones_menu = [
    "Nueva Solicitud"
]

if clave_admin == "BordaAdmin2026*":

    opciones_menu.append(
        "Consultas"
    )

    opciones_menu.append(
        "📊 Reportes"
    )

    opciones_menu.append(
        "Catálogo de Bordados"
    )


elif clave_admin:

    st.sidebar.error(
        "❌ Contraseña inválida"
    )

pagina = st.sidebar.selectbox(
    "Menú Principal",
    opciones_menu
)

if pagina == "📊 Reportes":

    st.title(
        "📊 Reportes Gerenciales"
    )

    df = obtener_ordenes()

    df_anulados = df[
        df["status"] == "Anulado"
    ]

    df = df[
        df["status"] != "Anulado"
    ]

    col1, col2, col3 = st.columns(3)

    with col1:

        st.metric(
            "📦 Total Pedidos",
            len(df)
        )

    with col2:

        st.metric(
            "💵 Saldo Pendiente",
            f"${df['saldo_pendiente'].sum():.2f}"
        )

    with col3:

        st.metric(
            "🟨 Recibidos",
            len(
                df[
                    df["status"] == "Recibido"
                ]
            )
        )

    col1, col2, col3 = st.columns(3)

    with col1:

        st.metric(
            "🔵 En Producción",
            len(
                df[
                    df["status"] == "En Producción"
                ]
            )
        )

    with col2:

        st.metric(
            "🟢 Listos para Entrega",
            len(
                df[
                    df["status"] == "Listo para Entrega"
                ]
            )
        )

    with col3:

        st.metric(
            "❌ Anulados",
            len(df_anulados)
        )


if pagina == "Catálogo de Bordados":

    st.title("🗂️ Administración de Catálogos")

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "⚙️ Configuración",
        "🏫 Colegios",
        "🚚 Delivery",
        "📦 Prendas",
        "🏷️ Marcas",
        "📏 Tallas",
        "🎨 Colores",
        "💾 Respaldo"
    ])
    
    with tab1:

        st.subheader(
            "⚙️ Configuración General"
        )

        precio_nombre = st.number_input(
            "Precio Bordado de Nombre",
            min_value=0.0,
            value=float(
                obtener_parametro(
                    "precio_nombre"
                )
            ),
            step=0.50
        )

        dias_produccion = st.number_input(
            "Días de Producción",
            min_value=1,
            value=int(
                obtener_parametro(
                    "dias_produccion"
                ) or 3
            ),
            step=1
        )

        if st.button(
            "💾 Guardar Configuración"
        ):

            guardar_parametro(
                "precio_nombre",
                precio_nombre
            )

            guardar_parametro(
                "dias_produccion",
                dias_produccion
            )

            st.success(
                "✅ Configuración actualizada"
            )    
    

    with tab2:

        st.subheader(
            "📚 Agregar Colegio"
        )
        df_colegios = obtener_colegios()

        nombre_colegio = st.text_input(
            "Nombre del Colegio"
        )
        precio_colegio = st.number_input(
            "Precio Bordado Colegio",
            min_value=0.0,
            step=0.50
        )

        if st.button(
            "💾 Guardar Colegio"
        ):

            guardar_colegio(
                nombre_colegio,
                precio_colegio
            )

            st.success(
                "✅ Colegio guardado"
            )

        st.subheader(
            "📚 Colegios Registrados"
        )
        st.dataframe(
            df_colegios,
            use_container_width=True
        )        

    with tab3:

        st.subheader(
            "🚚 Agregar Zona Delivery"
        )

        nombre_zona = st.text_input(
            "Nombre de la Zona"
        )

        costo_zona = st.number_input(
            "Costo Delivery",
            min_value=0.0,
            step=1.0
        )

        if st.button(
            "💾 Guardar Zona Delivery"
        ):

            guardar_zona_delivery(
                nombre_zona,
                costo_zona
            )

            st.success(
                "✅ Zona Delivery guardada"
            )

        df_zonas = obtener_zonas_delivery()

        st.subheader(
            "🚚 Zonas Delivery Registradas"
        )
        st.dataframe(
            df_zonas,
            use_container_width=True
        )

    with tab4:

        st.subheader(
            "📦 Agregar Tipo de Prenda"
        )

        nombre_tipo_prenda = st.text_input(
            "Nombre del Tipo de Prenda"
        )

        if st.button(
            "💾 Guardar Tipo de Prenda"
        ):

            guardar_tipo_prenda(
                nombre_tipo_prenda
            )

            st.success(
                "✅ Tipo de Prenda guardado"
            )

        st.subheader(
            "📦 Tipos de Prenda Registrados"
        )
        df_tipos_prenda = obtener_tipos_prenda()

        st.dataframe(
            df_tipos_prenda,
            use_container_width=True
        )        
        st.divider()
    with tab5:

        st.subheader(
            "🏷️ Agregar Marca"
        )

        nombre_marca = st.text_input(
            "Nombre de la Marca"
        )

        if st.button(
            "💾 Guardar Marca"
        ):

            guardar_marca(
                nombre_marca
            )

            st.success(
                "✅ Marca guardada"
            )

        st.subheader(
            "🏷️ Marcas Registradas"
        )
        df_marcas = obtener_marcas()

        st.dataframe(
            df_marcas,
            use_container_width=True
        )
        st.divider()

    with tab6:

        st.subheader(
            "📏 Agregar Talla"
        )

        nombre_talla = st.text_input(
            "Nombre de la Talla"
        )

        if st.button(
            "💾 Guardar Talla"
        ):

            guardar_talla(
                nombre_talla
            )

            st.success(
                "✅ Talla guardada"
            )

        st.subheader(
            "📏 Tallas Registradas"
        )

        df_tallas = obtener_tallas()

        st.dataframe(
            df_tallas,
            use_container_width=True
        )

        st.divider()

    with tab7:

        st.subheader(
            "🎨 Agregar Color"
        )

        nombre_color = st.text_input(
            "Nombre del Color"
        )

        if st.button(
            "💾 Guardar Color"
        ):

            guardar_color(
                nombre_color
            )

            st.success(
                "✅ Color guardado"
            )

        st.subheader(
            "🎨 Colores Registrados"
        )

        df_colores = obtener_colores()

        st.dataframe(
            df_colores,
            use_container_width=True
        )
                     
    with tab8:

        st.subheader(
            "💾 Respaldo Base de Datos"
        )

        st.info(
            "Descarga una copia completa de la base de datos actual de Bordaclick."
        )

        with open(
            "bordaclick.db",
            "rb"
        ) as archivo:

            st.download_button(
                label="📥 Descargar Base de Datos",
                data=archivo,
                file_name="bordaclick_backup.db",
                mime="application/octet-stream"
            )
        
def generar_pdf_orden(
    orden,
    detalle_orden
):

    nombre_pdf = f"Pedido_{int(orden['id']):04d}.pdf"

    doc = SimpleDocTemplate(nombre_pdf)

    estilos = getSampleStyleSheet()

    elementos = []

    logo = Image(
        "Logo Bordaclick.JPG",
        width=150,
        height=80
    )

    elementos.append(
        logo
    )

    elementos.append(
        Spacer(1, 10)
    )

    elementos.append(
        Paragraph(
            "ORDEN DE SERVICIO",
            estilos["Heading1"]
        )
    )

    elementos.append(
        Spacer(1, 10)
    )

    elementos.append(
        Spacer(1, 8)
    )

    elementos.append(
        Paragraph(
            f"Pedido #{int(orden['id']):04d}",
            estilos["Heading3"]
        )
    )

    tabla_estado = Table(
        [
            ["Estado", "Fecha Entrega"],
            [
                str(orden["status"]),
                str(orden["fecha_entrega"])
            ]
        ],
        colWidths=[120, 180]
    )

    tabla_estado.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER")
        ])
    )

    elementos.append(
        tabla_estado
    )

    elementos.append(
        Spacer(1, 8)
    )
    elementos.append(
        Paragraph(
            " ",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            "DATOS DEL CLIENTE",
            estilos["Heading3"]
        )
    )


    elementos.append(
        Paragraph(
            f"Nombre: {orden['nombre']}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Telefono: {orden['telefono']}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Correo: {orden['correo']}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Colegio: {orden['colegio']}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            " ",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            "DATOS DE PRODUCCION",
            estilos["Heading3"]
        )
    )
  

    elementos.append(
        Paragraph(
            f"Tipo de Logo: {orden['tipo_logo']}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Cantidad Total: {orden['cantidad_total']}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Nombre Bordado: {orden['nombre_bordado']}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Cantidad con Nombre: {orden['cantidad_nombre']}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            " ",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            "RESUMEN FINANCIERO",
            estilos["Heading3"]
        )
    )

    total_general = (
        orden["subtotal_bordado"]
        + orden["subtotal_nombres"]
        + orden["delivery_costo"]
    )

    tabla_financiera = Table([
        ["Concepto", "Monto"],
        ["Precio Bordado", f"${orden['precio_bordado']:.2f}"],
        ["Subtotal Bordado", f"${orden['subtotal_bordado']:.2f}"],
        ["Subtotal Nombres", f"${orden['subtotal_nombres']:.2f}"],
        ["Delivery", f"${orden['delivery_costo']:.2f}"],
        ["Total General", f"${total_general:.2f}"],
        ["Abono", f"${orden['abono']:.2f}"],
        ["Saldo Pendiente", f"${orden['saldo_pendiente']:.2f}"]
    ])

    tabla_financiera.setStyle(
        TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),

            ("BACKGROUND", (0, -1), (-1, -1), colors.yellow),

            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),

            ("ALIGN", (1, 1), (1, -1), "RIGHT")

        ])
    )

    elementos.append(
        tabla_financiera
    )

    total_general = (
        orden["subtotal_bordado"]
        + orden["subtotal_nombres"]
        + orden["delivery_costo"]
    )

    elementos.append(
        Paragraph(
            " ",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            "DESGLOSE DE PRENDAS",
            estilos["Heading3"]
        )
    )

    datos_tabla = [
        [
            "Tipo Prenda",
            "Talla",
            "Marca",
            "Color",
            "Cantidad"
        ]
    ]

    for _, fila in detalle_orden.iterrows():

        datos_tabla.append(
            [
                str(fila["Tipo Prenda"]),
                str(fila["Talla"]),
                str(fila["Marca"]),
                str(fila["Color"]),
                str(fila["Cantidad"])
            ]
        )

    tabla = Table(
        datos_tabla,
        colWidths=[90, 40, 90, 90, 50]
    )

    tabla.setStyle(
        TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("ALIGN", (1, 1), (-1, -1), "CENTER"),

            ("VALIGN", (0, 0), (-1, -1), "MIDDLE")

        ])
    )

    elementos.append(tabla)
    
    elementos.append(
        Spacer(1, 15)
    )

    elementos.append(
        Paragraph(
            "Bordaclick Diseños",
            estilos["Heading3"]
        )
    )

    elementos.append(
        Paragraph(
            "Sistema de Gestión de Bordados Escolares",
            estilos["Normal"]
        )
    )

    doc.build(elementos)

    return nombre_pdf


def generar_excel_orden(
    orden,
    detalle_orden
):


    nombre_excel = f"Pedido_{int(orden['id']):04d}.xlsx"

    wb = Workbook()

    ws = wb.active
    
    titulo = Font(
    bold=True
    )
    titulo_blanco = Font(
    bold=True,
    color="FFFFFF"
    )

    fondo_azul = PatternFill(
        fill_type="solid",
        start_color="ADD8E6"
    )
        
    borde = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
    )
    fondo_gris = PatternFill(
    fill_type="solid",
    start_color="D9D9D9",
    end_color="D9D9D9"
    )

    
    ws["A1"].font = titulo_blanco
    ws["A2"].font = titulo_blanco

    ws["A1"].fill = fondo_azul
    ws["A2"].fill = fondo_azul   
     
    ws.title = "Orden_Servicio"
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")

    ws["A1"] = "BORDACLICK"
    ws["A2"] = "ORDEN DE SERVICIO"

    ws["A4"] = "Pedido"
    ws["B4"] = f"#{int(orden['id']):04d}"

    ws["A5"] = "Estado"
    ws["B5"] = orden["status"]

    ws["A6"] = "Fecha Entrega"
    ws["B6"] = str(orden["fecha_entrega"])
    ws["A8"] = "DATOS DEL CLIENTE"

    ws["A9"] = "Nombre"
    ws["B9"] = orden["nombre"]

    ws["A10"] = "Telefono"
    ws["B10"] = orden["telefono"]

    ws["A11"] = "Correo"
    ws["B11"] = orden["correo"]

    ws["A12"] = "Colegio"
    ws["B12"] = orden["colegio"]
    ws["A14"] = "DATOS DE PRODUCCION"

    ws["A15"] = "Tipo Logo"
    ws["B15"] = orden["tipo_logo"]

    ws["A16"] = "Cantidad Total"
    ws["B16"] = orden["cantidad_total"]

    ws["A17"] = "Nombre Bordado"
    ws["B17"] = orden["nombre_bordado"]

    ws["A18"] = "Cantidad con Nombre"
    ws["B18"] = orden["cantidad_nombre"]
    ws["A20"] = "RESUMEN FINANCIERO"

    ws["A21"] = "Subtotal Bordado"
    ws["B21"] = orden["subtotal_bordado"]

    ws["A22"] = "Subtotal Nombres"
    ws["B22"] = orden["subtotal_nombres"]

    ws["A23"] = "Delivery"
    ws["B23"] = orden["delivery_costo"]

    ws["A24"] = "Abono"
    ws["B24"] = orden["abono"]

    ws["A25"] = "Saldo Pendiente"
    ws["B25"] = orden["saldo_pendiente"]
    ws["A27"] = "DESGLOSE DE PRENDAS"
    
    ws["A1"].font = titulo
    ws["A2"].font = titulo

    ws["A8"].font = titulo

    ws["A14"].font = titulo

    ws["A20"].font = titulo

    ws["A27"].font = titulo

    ws["A28"] = "Tipo Prenda"
    ws["B28"] = "Talla"
    ws["C28"] = "Marca"
    ws["D28"] = "Color"
    ws["E28"] = "Cantidad"
    fila_excel = 29

    for _, fila in detalle_orden.iterrows():

        ws[f"A{fila_excel}"] = fila["Tipo Prenda"]
        ws[f"B{fila_excel}"] = fila["Talla"]
        ws[f"C{fila_excel}"] = fila["Marca"]
        ws[f"D{fila_excel}"] = fila["Color"]
        ws[f"E{fila_excel}"] = fila["Cantidad"]

        fila_excel += 1
    encabezado_tabla = Font(
        bold=True
    )

    ws["A28"].font = encabezado_tabla
    ws["B28"].font = encabezado_tabla
    ws["C28"].font = encabezado_tabla
    ws["D28"].font = encabezado_tabla
    ws["E28"].font = encabezado_tabla
    ws["A28"].fill = fondo_gris
    ws["B28"].fill = fondo_gris
    ws["C28"].fill = fondo_gris
    ws["D28"].fill = fondo_gris
    ws["E28"].fill = fondo_gris

 
    for fila in ws.iter_rows(
        min_row=28,
        max_row=fila_excel - 1,
        min_col=1,
        max_col=5
    ):

        for celda in fila:

            celda.border = borde
    
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15    

    wb.save(nombre_excel)

    return nombre_excel
if pagina == "Consultas":

    st.title(
        "📋 Consulta de Órdenes"
    )

    pendientes = contar_pedidos_pendientes()

    if pendientes > 0:

        st.warning(
            f"🔔 Tienes {pendientes} pedidos nuevos por revisar"
        )

    else:

        st.success(
            "✅ No hay pedidos nuevos pendientes"
        )

    df_ordenes = obtener_ordenes()

    def colorear_estado(fila):

        estado = fila["status"]

        if estado == "Recibido":
            return ["background-color: #fff3cd"] * len(fila)

        elif estado == "En Producción":
            return ["background-color: #cfe2ff"] * len(fila)

        elif estado == "Listo para Entrega":
            return ["background-color: #d1e7dd"] * len(fila)

        return [""] * len(fila)

    st.dataframe(
        df_ordenes.style
        .format({
            "saldo_pendiente": "${:.2f}"
        })
        .apply(
            colorear_estado,
            axis=1
        ),
        use_container_width=True
    )

    pedido_id = st.selectbox(
        "Seleccione un pedido",
        df_ordenes["id"].tolist()
    )

    st.info(
        f"📦 Pedido seleccionado: #{pedido_id:04d}"
    )

    pedido = obtener_orden_por_id(
        pedido_id
    )

    with st.expander(
        "👤 Cliente",
        expanded=True
    ):

        st.write(
            f"**Nombre:** {pedido['nombre']}"
        )

        st.write(
            f"**Teléfono:** {pedido['telefono']}"
        )

        st.write(
            f"**Correo:** {pedido['correo']}"
        )

        if "colegio" in pedido:
            st.write(
                f"**Colegio:** {pedido['colegio']}"
            )
    with st.expander(
        "🏭 Producción"
    ):

        if st.session_state.get(
            "estado_actualizado",
            False
        ):

            st.success(
                "✅ Estado actualizado correctamente"
            )

            st.session_state["estado_actualizado"] = False

        st.write(
            f"**Estado actual:** {pedido['status']}"
        )

        estados = [
            "Recibido",
            "En Producción",
            "Listo para Entrega",
            "Anulado"
        ]

        indice_estado = (
            estados.index(pedido["status"])
            if pedido["status"] in estados
            else 0
        )

        nuevo_estado = st.selectbox(
            "Cambiar estado",
            estados,
            index=indice_estado
        )

        if st.button(
            "🔄 Actualizar Estado"
        ):

            actualizar_status_orden(
                pedido_id,
                nuevo_estado
            )

            if nuevo_estado in [
                "En Producción",
                "Listo para Entrega"
            ]:

                enviar_notificacion_estado(
                    pedido["correo"],
                    pedido["nombre"],
                    pedido["id"],
                    pedido["fecha_entrega"],
                    nuevo_estado,
                    pedido["delivery"]
                )

            st.session_state[
                "estado_actualizado"
            ] = True

            st.rerun()

        if "fecha_entrega" in pedido:
            st.write(
                f"**Fecha Entrega:** {pedido['fecha_entrega']}"
            )

        if "delivery" in pedido:
            st.write(
                f"**Delivery:** {pedido['delivery']}"
            )

        if "zona_delivery" in pedido:
            st.write(
                f"**Zona Delivery:** {pedido['zona_delivery']}"
            )
    with st.expander(
        "💰 Pagos"
    ):
            pass
    with st.expander(
        "💰 Resumen Financiero"
    ):

        col1, col2, col3 = st.columns(3)

        with col1:

            st.metric(
                "Abono",
                f"${pedido.get('abono',0):.2f}"
            )

        with col2:

            st.metric(
                "Saldo",
                f"${pedido.get('saldo_pendiente',0):.2f}"
            )

        with col3:

            total = (
                pedido.get("abono",0)
                +
                pedido.get("saldo_pendiente",0)
            )

            st.metric(
                "Total",
                f"${total:.2f}"
            )

    with st.expander(
        "👕 Prendas"
    ):

        detalle_orden = obtener_detalle_orden(
            pedido_id
        )

        st.dataframe(
            detalle_orden,
            use_container_width=True
        )
    with st.expander(
        "📄 Documentos"
    ):

        if st.button(
            "📄 Generar PDF"
        ):

            detalle_orden = obtener_detalle_orden(
                pedido_id
            )

            pdf_file = generar_pdf_orden(
                pedido,
                detalle_orden
            )

            st.success(
                "✅ PDF generado"
            )

            with open(
                pdf_file,
                "rb"
            ) as archivo:

                st.download_button(
                    "📥 Descargar PDF",
                    archivo,
                    file_name=pdf_file,
                    mime="application/pdf"
                )

        if st.button(
            "📊 Exportar Orden Excel"
        ):

            detalle_orden = obtener_detalle_orden(
                pedido_id
            )

            excel_file = generar_excel_orden(
                pedido,
                detalle_orden
            )

            with open(
                excel_file,
                "rb"
            ) as archivo:

                st.download_button(
                    "📥 Descargar Excel",
                    archivo,
                    file_name=excel_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
    with st.expander(
        "📧 Comunicaciones"
    ):

        st.info(
            f"Correo registrado: {pedido['correo']}"
        )

        if st.button(
            "📧 Reenviar Correo"
        ):

            detalle_orden = obtener_detalle_orden(
                pedido_id
            )

            pdf_file = generar_pdf_orden(
                pedido,
                detalle_orden
            )

            enviar_pdf_por_correo(
                pedido["correo"],
                pedido["nombre"],
                pedido["id"],
                pedido["fecha_entrega"],
                pdf_file
            )

            st.success(
                "✅ Correo enviado correctamente"
            )
                  
    with st.expander(
        "⚙️ Acciones"
    ):

        col1, col2 = st.columns(2)

        with col1:

            if st.button(
                "✏️ Editar Pedido"
            ):
                st.session_state["modo_edicion"] = True

        with col2:

            if st.button(
                "❌ Anular Pedido"
            ):

                actualizar_status_orden(
                    pedido_id,
                    "Anulado"
                )

                st.success(
                    "✅ Pedido anulado correctamente"
                )

                st.rerun()

        if st.session_state.get(
            "modo_edicion",
            False
        ):

            st.divider()

            st.success(
                f"✅ Editando Pedido #{pedido_id:04d}"
            )

            nombre_edit = st.text_input(
                "Nombre",
                value=pedido["nombre"]
            )

            telefono_edit = st.text_input(
                "Teléfono",
                value=pedido["telefono"]
            )

            correo_edit = st.text_input(
                "Correo",
                value=pedido["correo"]
            )

            if st.button(
                "💾 Guardar Cambios"
            ):
                st.success(
                    "✅ Cambios guardados"
                )        

  
if pagina == "Nueva Solicitud":

    st.subheader("Solicitud de Servicio de Bordado")

    st.header("Datos del Cliente")
    nombre = st.text_input(
        "Nombre y Apellido",
        key="nombre"
    )
    
    telefono = st.text_input(
        "Telefono / WhatsApp",
        key="telefono"
    )

    correo = st.text_input(
        "Correo Electronico",
        key="correo"
    )


    if telefono:
        if not re.match(r'^\d{10,}$', telefono):
           st.error("El teléfono debe tener solo números y mínimo 10 dígitos.")

    if correo:
        if not re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$', correo):
            st.error("Correo electrónico inválido.")

    
    df_colegios = obtener_colegios()

    lista_colegios = (
        ["Seleccione un colegio..."]
        +
        df_colegios["nombre"]
        .dropna()
        .tolist()
    )   
    
    colegio = st.selectbox(
        "Colegio",
        lista_colegios
    )
    precio_bordado = obtener_precio_colegio(
        colegio
    )
    

    st.info(
        f"💰 Precio del bordado: ${precio_bordado:.2f}"
    )

    st.header("Especificaciones Generales")

    logo = st.radio(
        "Tipo de logo",
        [
            "Logo de diario",
            "Logo de deporte",
            "Logo de preescolar"
        ]
    )


    dias_produccion = int(
    obtener_parametro(
        "dias_produccion"
    ) or 3
    )

    fecha_entrega = (
        date.today() +
        timedelta(days=dias_produccion)
    )
    st.info(
    f"📅 Fecha estimada de entrega: {fecha_entrega.strftime('%d/%m/%Y')}"
    )

    import pandas as pd

    st.header("Desglose de Prendas") 

    lista_tipos_prenda = (
        obtener_tipos_prenda()["nombre"]
        .dropna()
        .tolist()
    )

    lista_tallas = (
        obtener_tallas()["nombre"]
        .dropna()
        .tolist()
    )

    lista_marcas = (
        obtener_marcas()["nombre"]
        .dropna()
        .tolist()
    )

    lista_colores = (
        obtener_colores()["nombre"]
        .dropna()
        .tolist()
    )


    columnas = {
        "Tipo Prenda": st.column_config.SelectboxColumn(
            options=lista_tipos_prenda
        ),

        "Talla": st.column_config.SelectboxColumn(
            options=lista_tallas
        ),

        "Marca": st.column_config.SelectboxColumn(
            options=lista_marcas
        ),

        "Color": st.column_config.SelectboxColumn(
            options=lista_colores
        ),

        "Cantidad": st.column_config.NumberColumn(
            min_value=0,
            step=1
        )
    }

    df = st.data_editor(
        pd.DataFrame(
            columns=[
                "Tipo Prenda",
                "Talla",
                "Marca",
                "Color",
                "Cantidad"
            ]
        ),
        num_rows="dynamic",
        column_config=columnas,
        use_container_width=True
    )
    suma_prendas = 0


    if not df.empty:

        if "Cantidad" in df.columns:

            suma_prendas = df["Cantidad"].fillna(0).sum()

    cantidad_total = int(suma_prendas)
    

    if cantidad_total >= 6:

        precio_bordado = max(
            0,
            precio_bordado - 0.50
        )

        st.success(
            "🎉 Promoción aplicada: -$0.50 por bordado por cantidad (6 o más prendas)"
        )

    subtotal_bordado = (
        cantidad_total *
        precio_bordado
    )

    subtotal_bordado = (
        cantidad_total *
        precio_bordado
    )

    col1, col2 = st.columns(2)

    with col1:

        st.metric(
            "Total de Prendas",
            cantidad_total
        )

    with col2:

        st.metric(
            "Subtotal Bordado",
            f"${subtotal_bordado:.2f}"
        )

    cantidades_correctas = cantidad_total > 0


    nombre_bordado = ""

    cantidad_nombre = 0

    delivery = "No"

    zona_delivery = ""

    st.header("Bordado de Nombre")

    bordar_nombre = st.radio(
        "¿Desea bordar un nombre en sus prendas?",
        ["Sí", "No"]
    )

    nombre_bordado = ""
    cantidad_nombre = 0

    precio_nombre = obtener_parametro(
    "precio_nombre"
    )
    subtotal_nombres = 0
    cantidad_nombre = 0
    nombre_bordado = ""  

    if bordar_nombre == "Sí":

        nombre_bordado = st.text_area(
            "Indicar el nombre"
        )

        cantidad_nombre = st.number_input(
            "Cantidad de prendas con el nombre bordado",
            min_value=0,
            max_value=int(cantidad_total),
            step=1
        )

        subtotal_nombres = (
        cantidad_nombre *
        precio_nombre
        )

        st.info(
            f"💰 Precio por nombre: ${precio_nombre:.2f}"
            )

        st.metric(
            "Subtotal Bordado Nombre",
                f"${subtotal_nombres:.2f}"
                )

        if cantidad_nombre > cantidad_total:

            st.error(
                f"No puede bordar nombres en más de {cantidad_total} prendas."
            )

    
    st.header("Delivery")

    delivery = st.radio(
        "¿Desea servicio delivery?",
        [
            "Sí (con costo adicional)",
            "No"
        ]
    )

    zona_delivery = ""
    delivery_costo = 0

    if delivery == "Sí (con costo adicional)":

        df_zonas = obtener_zonas_delivery()

        lista_zonas = (
            ["Seleccione zona de delivery..."]
            +
            df_zonas["nombre"]
            .dropna()
            .tolist()
        )

        zona_delivery = st.selectbox(
            "Zona de entrega",
            lista_zonas
        )

        if zona_delivery != "Seleccione zona de delivery...":
            if st.session_state.get(
                "delivery_ya_cobrado",
                False
            ):

                delivery_costo = 0

                st.info(
                    "🚚 Delivery ya cobrado en una orden anterior."
                )

            else:

                delivery_costo = obtener_costo_delivery(
                    zona_delivery
                )


            st.info(
                f"🚚 Costo Delivery: ${delivery_costo:.2f}"
            )

        st.metric(
            "Subtotal Delivery",
            f"${delivery_costo:.2f}"
        )

    else:

        zona_delivery = ""
        delivery_costo = 0      

    st.header("Información de Pago")

    abono = 0.0

    saldo_pendiente = (
        subtotal_bordado +
        subtotal_nombres +
        delivery_costo -
        abono
    )

    st.metric(
        "Saldo Pendiente",
        f"${saldo_pendiente:.2f}"
    )
    if "solicitud_guardada" not in st.session_state:
        st.session_state["solicitud_guardada"] = False

    if st.button(
        "Guardar Solicitud",
        disabled=(
            not cantidades_correctas
            or st.session_state["solicitud_guardada"]
        )
    ):

        # === VALIDACIONES ===

        if colegio == "Seleccione un colegio...":

            st.error(
                "Debe seleccionar un colegio"
            )

            st.stop()

        if (
            delivery == "Sí (con costo adicional)"
            and
            zona_delivery == "Seleccione zona de delivery..."
        ):

            st.error(
                "Debe seleccionar una zona de delivery"
            )

            st.stop()

        status = "Recibido"

        orden_id = guardar_orden(
            nombre,
            telefono,
            correo,
            colegio,
            cantidad_total,
            logo,
            nombre_bordado,
            cantidad_nombre,
            delivery,
            zona_delivery,
            fecha_entrega,
            precio_bordado,
            subtotal_bordado,
            subtotal_nombres,
            delivery_costo,
            abono,
            saldo_pendiente,
            status
        )
        for _, fila in df.iterrows():

            guardar_detalle(
                orden_id,
                fila["Tipo Prenda"],
                fila["Talla"],
                fila["Marca"],
                fila["Color"],
                int(fila["Cantidad"])
            )

        detalle_orden = obtener_detalle_orden(
            orden_id
        )

        orden_generada = obtener_orden_por_id(
            orden_id
        )

        pdf_file = generar_pdf_orden(
            orden_generada,
            detalle_orden
        )

        enviar_pdf_por_correo(
            correo,
            nombre,
            orden_id,
            fecha_entrega,
            pdf_file
        )

        st.session_state["solicitud_guardada"] = True

        st.rerun()


    if st.session_state["solicitud_guardada"]:

        st.success(
            "✅ Solicitud guardada correctamente"
        )

        if st.button("➕ Nueva Solicitud"):

            st.session_state["solicitud_guardada"] = False

            if "nombre" in st.session_state:
                del st.session_state["nombre"]

            if "telefono" in st.session_state:
                del st.session_state["telefono"]

            if "correo" in st.session_state:
                del st.session_state["correo"]

            st.rerun()






    

