import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


def _safe_float(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def generar_pdf_orden(orden, detalle_orden):
    pedido_id = int(_safe_float(orden.get("id")))
    nombre_pdf = f"Pedido_{pedido_id:04d}.pdf"

    doc = SimpleDocTemplate(
        nombre_pdf,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    estilos = getSampleStyleSheet()
    elementos = []

    try:
        logo = Image("Logo Bordaclick.JPG", width=140, height=70)
        elementos.append(logo)
        elementos.append(Spacer(1, 10))
    except Exception:
        pass

    elementos.append(Paragraph("ORDEN DE SERVICIO", estilos["Heading1"]))
    elementos.append(Spacer(1, 6))
    elementos.append(Paragraph(f"Pedido #{pedido_id:04d}", estilos["Heading3"]))

    tabla_estado = Table(
        [
            ["Estado", "Fecha Entrega"],
            [str(orden.get("status", "")), str(orden.get("fecha_entrega", ""))]
        ],
        colWidths=[140, 180]
    )
    tabla_estado.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A365D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER")
        ])
    )
    elementos.append(tabla_estado)
    elementos.append(Spacer(1, 12))

    elementos.append(Paragraph("<b>DATOS DEL CLIENTE</b>", estilos["Heading3"]))
    elementos.append(Paragraph(f"<b>Nombre:</b> {orden.get('nombre', '')}", estilos["Normal"]))
    elementos.append(Paragraph(f"<b>Teléfono:</b> {orden.get('telefono', '')}", estilos["Normal"]))
    elementos.append(Paragraph(f"<b>Correo:</b> {orden.get('correo', '')}", estilos["Normal"]))
    elementos.append(Paragraph(f"<b>Colegio Principal:</b> {orden.get('colegio', '')}", estilos["Normal"]))
    elementos.append(Spacer(1, 10))

    elementos.append(Paragraph("<b>DATOS DE PRODUCCIÓN</b>", estilos["Heading3"]))
    elementos.append(Paragraph(f"<b>Tipo de Logo:</b> {orden.get('tipo_logo', '')}", estilos["Normal"]))
    elementos.append(Paragraph(f"<b>Cantidad Total:</b> {orden.get('cantidad_total', 0)}", estilos["Normal"]))
    elementos.append(Paragraph(f"<b>Bordado Nombre:</b> {orden.get('nombre_bordado', 'N/A')}", estilos["Normal"]))
    elementos.append(Paragraph(f"<b>Prendas con Nombre:</b> {orden.get('cantidad_nombre', 0)}", estilos["Normal"]))
    elementos.append(Spacer(1, 10))

    sub_bordado = _safe_float(orden.get("subtotal_bordado"))
    sub_nombres = _safe_float(orden.get("subtotal_nombres"))
    del_costo = _safe_float(orden.get("delivery_costo"))
    p_bordado = _safe_float(orden.get("precio_bordado"))
    abono = _safe_float(orden.get("abono"))
    saldo = _safe_float(orden.get("saldo_pendiente"))
    total_general = sub_bordado + sub_nombres + del_costo

    tabla_financiera = Table([
        ["Concepto", "Monto"],
        ["Precio Bordado Unitario", f"${p_bordado:.2f}"],
        ["Subtotal Bordado", f"${sub_bordado:.2f}"],
        ["Subtotal Nombres", f"${sub_nombres:.2f}"],
        ["Delivery", f"${del_costo:.2f}"],
        ["Total General", f"${total_general:.2f}"],
        ["Abono", f"${abono:.2f}"],
        ["Saldo Pendiente", f"${saldo:.2f}"]
    ], colWidths=[200, 120])

    tabla_financiera.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A365D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF3BF")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT")
        ])
    )
    elementos.append(Paragraph("<b>RESUMEN FINANCIERO</b>", estilos["Heading3"]))
    elementos.append(tabla_financiera)
    elementos.append(Spacer(1, 15))

    elementos.append(Paragraph("<b>DESGLOSE DE PRENDAS</b>", estilos["Heading3"]))

    if not detalle_orden.empty and "Colegio" in detalle_orden.columns:
        for colegio in detalle_orden["Colegio"].unique():
            elementos.append(Paragraph(f"🏫 <b>{colegio}</b>", estilos["Heading4"]))

            datos_tabla = [["Tipo Prenda", "Talla", "Marca", "Color", "Cantidad"]]
            df_colegio = detalle_orden[detalle_orden["Colegio"] == colegio]

            for _, fila in df_colegio.iterrows():
                datos_tabla.append([
                    str(fila.get("Tipo Prenda", "")),
                    str(fila.get("Talla", "")),
                    str(fila.get("Marca", "")),
                    str(fila.get("Color", "")),
                    str(fila.get("Cantidad", ""))
                ])

            tabla_prendas = Table(datos_tabla, colWidths=[120, 50, 100, 100, 60])
            tabla_prendas.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B6CB0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE")
                ])
            )
            elementos.append(tabla_prendas)
            elementos.append(Spacer(1, 10))

    elementos.append(Spacer(1, 15))
    elementos.append(Paragraph("<b>Bordaclick Diseños</b>", estilos["Heading3"]))
    elementos.append(Paragraph("Sistema de Gestión de Bordados Escolares", estilos["Normal"]))

    doc.build(elementos)
    return nombre_pdf


def generar_excel_orden(orden, detalle_orden):
    pedido_id = int(_safe_float(orden.get("id")))
    nombre_excel = f"Pedido_{pedido_id:04d}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Orden_Servicio"

    titulo_blanco = Font(bold=True, color="FFFFFF", size=12)
    subtitulo_bold = Font(bold=True, size=11)
    encabezado_tabla = Font(bold=True, color="FFFFFF")
    
    fondo_azul_oscuro = PatternFill(fill_type="solid", start_color="1A365D")
    fondo_azul_medio = PatternFill(fill_type="solid", start_color="2B6CB0")

    borde = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")

    ws["A1"] = "BORDACLICK"
    ws["A2"] = "ORDEN DE SERVICIO"

    for cell in [ws["A1"], ws["A2"]]:
        cell.font = titulo_blanco
        cell.fill = fondo_azul_oscuro
        cell.alignment = Alignment(horizontal="center", vertical="center")

    datos = [
        ("A4", "Pedido", f"#{pedido_id:04d}"),
        ("A5", "Estado", orden.get("status", "")),
        ("A6", "Fecha Entrega", str(orden.get("fecha_entrega", ""))),
        ("A8", "DATOS DEL CLIENTE", ""),
        ("A9", "Nombre", orden.get("nombre", "")),
        ("A10", "Teléfono", orden.get("telefono", "")),
        ("A11", "Correo", orden.get("correo", "")),
        ("A12", "Colegio", orden.get("colegio", "")),
        ("A14", "DATOS DE PRODUCCIÓN", ""),
        ("A15", "Tipo Logo", orden.get("tipo_logo", "")),
        ("A16", "Cantidad Total", orden.get("cantidad_total", 0)),
        ("A17", "Nombre Bordado", orden.get("nombre_bordado", "N/A")),
        ("A18", "Cantidad con Nombre", orden.get("cantidad_nombre", 0)),
        ("A20", "RESUMEN FINANCIERO", ""),
        ("A21", "Subtotal Bordado", _safe_float(orden.get("subtotal_bordado"))),
        ("A22", "Subtotal Nombres", _safe_float(orden.get("subtotal_nombres"))),
        ("A23", "Delivery", _safe_float(orden.get("delivery_costo"))),
        ("A24", "Abono", _safe_float(orden.get("abono"))),
        ("A25", "Saldo Pendiente", _safe_float(orden.get("saldo_pendiente")))
    ]

    for c_pos, label, val in datos:
        ws[c_pos] = label
        row_num = int(c_pos[1:])
        
        if label in ["DATOS DEL CLIENTE", "DATOS DE PRODUCCIÓN", "RESUMEN FINANCIERO"]:
            ws[c_pos].font = subtitulo_bold
        else:
            ws[f"B{row_num}"] = val
            ws[c_pos].font = Font(bold=True)
            if isinstance(val, float):
                ws[f"B{row_num}"].number_format = "$#,##0.00"

    fila_excel = 27
    ws[f"A{fila_excel}"] = "DESGLOSE DE PRENDAS"
    ws[f"A{fila_excel}"].font = subtitulo_bold
    fila_excel += 1

    inicio_desglose = fila_excel

    if not detalle_orden.empty and "Colegio" in detalle_orden.columns:
        for colegio in detalle_orden["Colegio"].unique():
            ws[f"A{fila_excel}"] = f"🏫 {colegio}"
            ws[f"A{fila_excel}"].font = subtitulo_bold
            fila_excel += 1

            headers = ["Tipo Prenda", "Talla", "Marca", "Color", "Cantidad"]
            cols = ["A", "B", "C", "D", "E"]

            for col_letter, h_text in zip(cols, headers):
                cell = ws[f"{col_letter}{fila_excel}"]
                cell.value = h_text
                cell.font = encabezado_tabla
                cell.fill = fondo_azul_medio
                cell.alignment = Alignment(horizontal="center")

            fila_excel += 1

            df_colegio = detalle_orden[detalle_orden["Colegio"] == colegio]
            for _, fila in df_colegio.iterrows():
                ws[f"A{fila_excel}"] = fila.get("Tipo Prenda", "")
                ws[f"B{fila_excel}"] = fila.get("Talla", "")
                ws[f"C{fila_excel}"] = fila.get("Marca", "")
                ws[f"D{fila_excel}"] = fila.get("Color", "")
                ws[f"E{fila_excel}"] = fila.get("Cantidad", 0)
                fila_excel += 1

            fila_excel += 1

    for fila in ws.iter_rows(min_row=inicio_desglose, max_row=fila_excel - 1, min_col=1, max_col=5):
        for celda in fila:
            if celda.value:
                celda.border = borde

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15

    wb.save(nombre_excel)
    return nombre_excel